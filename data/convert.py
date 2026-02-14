"""
One-time script to convert jobs.xlsx → jobs.csv with region column.

Each sheet in the xlsx represents a region. The output CSV has columns:
  region, programme_title,
  p1_site, p1_speciality, p2_site, p2_speciality, p3_site, p3_speciality,
  p4_site, p4_speciality, p4_description,
  p5_site, p5_speciality, p5_description,
  p6_site, p6_speciality, p6_description

Run:
  cd data && uv run python convert.py
"""

import csv
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "jobs.xlsx"
CSV_PATH = Path(__file__).parent / "jobs.csv"
PUBLIC_PATH = Path(__file__).parent / ".." / "web" / "public" / "jobs.csv"

# Column mappings — the xlsx uses slightly inconsistent naming
SITE_KEYS = ["Placement {n}: Site", "Placement {n} Site"]
SPEC_KEYS = [
    "Placement {n}: Specialty",
    "Placement {n}: Speciality",
    "Placement {n} Specialty",
]
DESC_KEYS = ["Placement {n}: Description", "Placement {n} Description"]


def find_col(header_map: dict[str, int], templates: list[str], n: int) -> int | None:
    for t in templates:
        key = t.replace("{n}", str(n))
        if key in header_map:
            return header_map[key]
    return None


def convert() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    csv_columns = ["region", "programme_title"]
    for i in range(1, 7):
        csv_columns.append(f"p{i}_site")
        csv_columns.append(f"p{i}_speciality")
        if i >= 4:
            csv_columns.append(f"p{i}_description")

    rows: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        headers = [str(h).strip() if h else "" for h in all_rows[0]]
        header_map = {h: idx for idx, h in enumerate(headers)}

        title_col = header_map.get("Programme Title")
        if title_col is None:
            continue

        for row in all_rows[1:]:
            title = row[title_col] if title_col < len(row) else None
            if not title or str(title).strip() == "":
                continue

            out: dict[str, str] = {
                "region": sheet_name,
                "programme_title": str(title).strip(),
            }

            for i in range(1, 7):
                site_idx = find_col(header_map, SITE_KEYS, i)
                spec_idx = find_col(header_map, SPEC_KEYS, i)
                desc_idx = find_col(header_map, DESC_KEYS, i)

                site = row[site_idx] if site_idx is not None and site_idx < len(row) else None
                spec = row[spec_idx] if spec_idx is not None and spec_idx < len(row) else None
                desc = row[desc_idx] if desc_idx is not None and desc_idx < len(row) else None

                out[f"p{i}_site"] = str(site).strip() if site else ""
                out[f"p{i}_speciality"] = str(spec).strip() if spec else ""
                if i >= 4:
                    out[f"p{i}_description"] = str(desc).strip() if desc else ""

            rows.append(out)

    wb.close()

    # Write to data/jobs.csv
    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_columns)
        writer.writeheader()
        writer.writerows(rows)

    # Also copy to web/public/ for the app
    PUBLIC_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(PUBLIC_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_columns)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} jobs to {CSV_PATH}")
    print(f"Copied to {PUBLIC_PATH}")


if __name__ == "__main__":
    convert()
